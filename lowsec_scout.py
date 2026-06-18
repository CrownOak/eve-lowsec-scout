#!/usr/bin/env python3
"""
CROWN & OAK CAPITAL - Lowsec Scout  (v3, file-only / no database)
================================================================
Finds SAFE, LOW-TRAFFIC mining systems and keeps a running record of them, so
you can spot genuinely quiet space instead of clicking through DOTLAN region by
region, and tell the difference between "quiet right now" and "actually quiet."

Data sources (all public, no API key, no auth):
  ESI  /universe/system_kills/   ship/pod/npc kills per system   (LAST HOUR only)
  ESI  /universe/system_jumps/   jumps (traffic) per system       (last hour)
  SDE  (Fuzzwork)                name, security, TRUESEC, region  (cached locally)
  zKill /api/systemID/<id>/      recent kill history per system   (shortlist only)

WHAT CHANGED IN v3 (vs the SQLite build)
  No database. Persistence is now two plain flat files, exactly as requested:
    * scout_history.csv  -> the running record. One snapshot per run is APPENDED
      here (append only, so it is fast and never rewrites the bulk). This is what
      builds a real multi-day danger history and fixes "ESI only shows one hour."
    * lowsec_scout.xlsx  -> the running, human-readable workbook. A "Shortlist"
      sheet is refreshed every run with the current ranked table, and a "RunLog"
      sheet gets one summary row appended per run. Open it any time in Excel.
  Also fixed: the Fuzzwork SDE path moved under /dump/latest/csv/, and that CSV
  now ships a UTF-8 BOM, both of which broke the old downloader.

WHAT EACH RUN DOES
  1. RECORDER  -> appends this hour's ESI snapshot to scout_history.csv. Run it on
     a schedule (Task Scheduler / cron) and it accumulates per-system history.
  2. SCORING   -> each system gets a SAFETY score (kills + traffic, weighting pod
     kills as a camp signal, blended with its recorded history once >=3 samples
     exist) and an ORE-POTENTIAL score (lower truesec = richer ore). Final rank =
     mostly safety, then ore. Use --space to scan lowsec / hisec / null.
  3. THREAT    -> the top-N shortlist is cross-checked against zKillboard for
     recent kill volume and how long since the last kill, extending the 1-hour ESI
     window to a multi-day picture so a between-victims gate camp cannot pose as
     "quiet."

IMPORTANT, READ THIS: no public API exposes the live ore ANOMALIES in a system.
That only exists in your in-game probe scanner. This tool ranks safe, low-traffic,
ore-SUITABLE candidate systems. You still warp in and scan to confirm the rocks.

USAGE
    python lowsec_scout.py                      # 25 best lowsec candidates, whole game
    python lowsec_scout.py --space hisec        # highsec mode
    python lowsec_scout.py --region Aridia --top 40
    python lowsec_scout.py --min-jumps 1        # drop totally dead 0-traffic systems
    python lowsec_scout.py --no-threat          # skip zKill lookups (faster)
    python lowsec_scout.py --out scout.csv      # also write a one-off CSV of the table
    python lowsec_scout.py --no-xlsx            # skip the workbook (pure CSV mode)
    python lowsec_scout.py --demo               # offline self-test, fake data + files

Schedule it (hourly) to build history. The included run.ps1 / run.bat do this.

Requires Python 3.8+. Standard library only EXCEPT openpyxl (for the .xlsx). If
openpyxl is not installed the tool still runs and keeps scout_history.csv; it just
skips the workbook. Live runs need open internet (ESI / Fuzzwork / zKill).
"""

import argparse, csv, io, os, sys, tempfile, time
import urllib.request, urllib.error
from datetime import datetime, timezone, timedelta

# Render cleanly on a Windows console (the default cp1252 mangles non-ASCII).
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ESI       = "https://esi.evetech.net/latest"
# Fuzzwork moved the CSV dumps under /csv/ and they now carry a UTF-8 BOM.
SDE_SYS   = "https://www.fuzzwork.co.uk/dump/latest/csv/mapSolarSystems.csv"
SDE_REG   = "https://www.fuzzwork.co.uk/dump/latest/csv/mapRegions.csv"
ZKILL     = "https://zkillboard.com/api"
# A descriptive UA with a real contact is required by zKill/CCP.
UA        = "CrownOakLowsecScout/3.0 (Crown & Oak Capital; salesmaxxllc@gmail.com)"
SYS_CACHE = "systems_cache.json"
HIST_CSV  = "scout_history.csv"
XLSX_PATH = "lowsec_scout.xlsx"
HIST_COLS = ["ts", "system_id", "ship", "pod", "npc", "jumps"]
ZKILL_DELAY = 1.1          # be polite to zKill (~1 req/sec)


# ----------------------------------------------------------------------------
# HTTP
# ----------------------------------------------------------------------------

def fetch(url, parse_json=True, timeout=60, encoding="utf-8"):
    import json
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    return json.loads(raw) if parse_json else raw.decode(encoding, "replace")


def post_webhook(url, payload, token=None, timeout=10):
    """POST a JSON payload to a (local) webhook. Returns (status, body)."""
    import json
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json", "User-Agent": UA}
    if token:
        headers["X-Webhook-Token"] = token
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.status, r.read().decode("utf-8", "replace")


# ----------------------------------------------------------------------------
# SYSTEM CACHE (name, security, truesec, region)  -- one-time SDE download
# ----------------------------------------------------------------------------

def build_system_cache():
    import json
    if os.path.exists(SYS_CACHE):
        with open(SYS_CACHE, encoding="utf-8") as f:
            return json.load(f)
    print("  First run: downloading EVE system map from Fuzzwork (one-time, ~couple MB)...")
    # utf-8-sig strips the BOM so the first CSV column parses as 'regionID', not '﻿"regionID"'.
    sys_csv = fetch(SDE_SYS, parse_json=False, encoding="utf-8-sig")
    reg_csv = fetch(SDE_REG, parse_json=False, encoding="utf-8-sig")
    regions = {row["regionID"]: row["regionName"] for row in csv.DictReader(io.StringIO(reg_csv))}
    cache = {}
    for row in csv.DictReader(io.StringIO(sys_csv)):
        rid = row["regionID"]
        try:
            if int(rid) >= 11000000:      # skip wormhole / abyssal (k-space only)
                continue
        except (ValueError, TypeError):
            continue
        try:
            true = float(row["security"])     # full-precision truesec
        except (ValueError, TypeError):
            continue
        cache[row["solarSystemID"]] = {
            "name": row["solarSystemName"],
            "sec": round(true, 2),        # rounded, for band + display
            "truesec": round(true, 4),    # precise, drives ore-potential
            "region_id": rid,
            "region": regions.get(rid, rid),
        }
    with open(SYS_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)
    print(f"  Cached {len(cache)} systems.\n")
    return cache


def in_band(sec, space):
    if space == "lowsec": return 0.05 <= sec < 0.45
    if space == "hisec":  return sec >= 0.45
    if space == "null":   return sec < 0.05
    return True  # "all"


# ----------------------------------------------------------------------------
# LIVE ACTIVITY (ESI, last hour)
# ----------------------------------------------------------------------------

def get_live_activity():
    kills = fetch(f"{ESI}/universe/system_kills/")
    jumps = fetch(f"{ESI}/universe/system_jumps/")
    k = {str(x["system_id"]): (x.get("ship_kills", 0), x.get("pod_kills", 0), x.get("npc_kills", 0)) for x in kills}
    j = {str(x["system_id"]): x.get("ship_jumps", 0) for x in jumps}
    return k, j


# ----------------------------------------------------------------------------
# RECORDER  (flat CSV, append-only) -- replaces the old SQLite snapshots table
# ----------------------------------------------------------------------------

def history_record(path, ts, kills, jumps):
    """Append this hour's snapshot to the running history CSV.
    Only rows with any activity are written, to keep the file lean (this mirrors
    the old db_record behaviour exactly). Returns the number of rows written."""
    rows, seen = [], set()
    for sid, (sk, pk, nk) in kills.items():
        rows.append((ts, sid, sk, pk, nk, jumps.get(sid, 0))); seen.add(sid)
    for sid, jp in jumps.items():
        if sid not in seen and jp:
            rows.append((ts, sid, 0, 0, 0, jp))
    new_file = not os.path.exists(path) or os.path.getsize(path) == 0
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(HIST_COLS)
        w.writerows(rows)
    return len(rows)


def history_load(path, days):
    """Per-system stats over the last <days>, computed by streaming the CSV.
    Returns ({sid: (avg_ship, avg_pod, samples, max_pvp)}, snapshot_count). The
    per-system contract is identical to the old db_history() so the scoring code
    did not change; snapshot_count is how many distinct runs fall in the window."""
    if not os.path.exists(path):
        return {}, 0
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    acc = {}  # sid -> [sum_ship, sum_pod, count, max_pvp]
    snaps = set()
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            ts = row.get("ts", "")
            if ts < cutoff:           # ISO-8601 UTC strings compare lexicographically
                continue
            sid = row.get("system_id")
            if not sid:
                continue
            try:
                ship = int(row.get("ship") or 0)
                pod  = int(row.get("pod") or 0)
            except ValueError:
                continue
            snaps.add(ts)
            a = acc.get(sid)
            if a is None:
                acc[sid] = [ship, pod, 1, ship + pod]
            else:
                a[0] += ship; a[1] += pod; a[2] += 1
                if ship + pod > a[3]:
                    a[3] = ship + pod
    return {sid: (s / n, p / n, n, mx) for sid, (s, p, n, mx) in acc.items()}, len(snaps)


def history_prune(path, retain_days):
    """Keep the running CSV from growing forever. If its OLDEST row is older than
    the retain window, rewrite the file atomically with only rows still in window.
    Append-only means the file is chronological, so we only act when needed."""
    if retain_days is None or not os.path.exists(path) or os.path.getsize(path) == 0:
        return 0
    cutoff = (datetime.now(timezone.utc) - timedelta(days=retain_days)).isoformat()
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return 0
        first = next(reader, None)
        if first is None or not first or first[0] >= cutoff:
            return 0  # nothing old enough to drop (or a blank first line); skip rewrite
    # A rewrite is warranted. Stream into a temp file in the same dir, then replace.
    dropped = 0
    d = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(suffix=".csv", dir=d)
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as out, \
             open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f); writer = csv.writer(out)
            writer.writerow(next(reader))  # header
            for row in reader:
                if row and row[0] >= cutoff:
                    writer.writerow(row)
                else:
                    dropped += 1
        os.replace(tmp, path)
    except OSError:
        # file held open (Excel/editor/concurrent run); skip this prune, retry next run.
        # The snapshot was already appended, so nothing is lost by leaving it un-pruned.
        try: os.remove(tmp)
        except OSError: pass
        return 0
    except Exception:
        try: os.remove(tmp)
        except OSError: pass
        raise
    return dropped


# ----------------------------------------------------------------------------
# SCORING  (unchanged from v2 -- the build spec says do not touch the logic)
# ----------------------------------------------------------------------------

def ore_potential(truesec):
    """Lower truesec -> richer ore / higher grades. 0.5 sec -> 0, -1.0 -> 100."""
    return max(0.0, min(100.0, (0.5 - truesec) / 1.5 * 100.0))

def danger_weighted(ship, pod):
    """Pods weighted double -- lots of pod kills signals smartbomb/gate camps."""
    return ship + 2.0 * pod

def safety_score(dw):
    return max(0, round(100 - 14 * dw))

def composite(safety, ore):
    """Safety-led (you asked for SAFE first), ore as the secondary pull."""
    return round(0.70 * safety + 0.30 * ore)


# ----------------------------------------------------------------------------
# THREAT INTEL  (zKill, shortlist only)
# ----------------------------------------------------------------------------

def zkill_recent(system_id):
    """Recent kill volume + age of the most recent kill for one system.
    Returns (kill_count, hours_since_last) or (None, None) on failure."""
    if not system_id:
        return (None, None)
    try:
        data = fetch(f"{ZKILL}/systemID/{system_id}/")
    except Exception:
        return (None, None)
    finally:
        time.sleep(ZKILL_DELAY)
    if not isinstance(data, list):
        return (0, None)
    count = len(data)
    last_h = None
    for entry in data[:1]:
        t = entry.get("killmail_time") if isinstance(entry, dict) else None
        if t:
            try:
                kt = datetime.fromisoformat(t.replace("Z", "+00:00"))
                last_h = round((datetime.now(timezone.utc) - kt).total_seconds() / 3600, 1)
            except ValueError:
                pass
    return (count, last_h)


# ----------------------------------------------------------------------------
# XLSX OUTPUT  (running workbook: Shortlist refreshed + RunLog appended)
# ----------------------------------------------------------------------------

SHORTLIST_COLS = [
    ("name", "SYSTEM", 16), ("sec", "SEC", 6), ("region", "REGION", 18),
    ("score", "SCORE", 7), ("safety", "SAFE", 6), ("ore", "ORE", 6),
    ("pvp1h", "PVP1h", 7), ("jumps", "JUMP", 6), ("hist_avg_pvp", "~AVG", 7),
    ("samples", "SAMPLES", 9), ("zk_recent", "ZK", 6), ("zk_last_h", "LASTh", 8),
    ("truesec", "TRUESEC", 9),
]

def _sheet_name(space, region):
    """Shortlist sheet name keyed by space (and region) so a lowsec hourly run never
    clobbers a manually-built hisec shortlist. Sanitised to Excel's rules + 31 chars."""
    base = f"Shortlist-{space}"
    if region:
        base += "-" + region
    for ch in r'\/?*[]:':
        base = base.replace(ch, "_")
    return base[:31]


def write_xlsx(path, rows, meta):
    """Rebuild the Shortlist sheet for this space/region and append one RunLog row.
    The save is atomic against a crash and tolerant of Excel holding the file open
    (it is NOT a concurrent-writer lock). Returns the path written, or None if
    openpyxl is unavailable."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (openpyxl not installed; skipping .xlsx. `pip install openpyxl` to enable it.)")
        return None

    wb = None
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception:
            wb = None  # corrupt / locked-on-read; start fresh
    if wb is None:
        wb = Workbook()
        wb.remove(wb.active)
    # one-time cleanup: drop the legacy single 'Shortlist' sheet from older versions
    if "Shortlist" in wb.sheetnames:
        wb.remove(wb["Shortlist"])

    # ---- Shortlist sheet for THIS space/region (rebuilt every run) ----
    sheet = _sheet_name(meta["space"], meta["region"])
    if sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    ws = wb.create_sheet(sheet, 0)   # newest view to the front

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1A1A1A")
    title_font = Font(bold=True, size=13, color="1A1A1A")
    link_font = Font(color="1F5FBF", underline="single")
    good = PatternFill("solid", fgColor="C9E7C9")   # safe / high score
    warn = PatternFill("solid", fgColor="F6E2A0")   # middling
    bad  = PatternFill("solid", fgColor="F0C0BE")   # risky

    ws["A1"] = "CROWN & OAK  .  Lowsec Scout"
    ws["A1"].font = title_font
    threat_note = ("zKill threat data included" if meta["threat"]
                   else "recorder snapshot, no zKill data (blank ZK does not mean zero kills)")
    ws["A2"] = (f"generated {meta['ts']}  .  space={meta['space']}  "
                f"region={meta['region'] or 'all'}  .  history {meta['history_days']}d "
                f"({meta['snapshots_in_window']} snapshots)  .  {len(rows)} systems  .  {threat_note}")
    ws["A2"].font = Font(italic=True, color="555555")

    hrow = 4
    for c, (_, label, width) in enumerate(SHORTLIST_COLS, start=1):
        cell = ws.cell(row=hrow, column=c, value=label)
        cell.font = head_font; cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    for i, r in enumerate(rows):
        rr = hrow + 1 + i
        vals = {
            "name": r["name"], "sec": r["sec"], "region": r["region"],
            "score": r["score"], "safety": r["safety"], "ore": r["ore"],
            "pvp1h": r["ship"] + r["pod"], "jumps": r["jumps"],
            "hist_avg_pvp": r["hist_avg_pvp"], "samples": r["samples"],
            "zk_recent": r["zk_recent"], "zk_last_h": r["zk_last_h"],
            "truesec": r["truesec"],
        }
        for c, (key, _, _) in enumerate(SHORTLIST_COLS, start=1):
            ws.cell(row=rr, column=c, value=vals[key])
        # DOTLAN map link on the system name so the miner can jump straight to it
        name_cell = ws.cell(row=rr, column=1)
        name_cell.hyperlink = "https://evemaps.dotlan.net/system/" + str(r["name"]).replace(" ", "_")
        name_cell.font = link_font
        # colour the SAFE cell as a quick visual cue
        safe_cell = ws.cell(row=rr, column=5)
        safe_cell.fill = good if r["safety"] >= 85 else warn if r["safety"] >= 50 else bad

    last_col = get_column_letter(len(SHORTLIST_COLS))
    if rows:
        ws.auto_filter.ref = f"A{hrow}:{last_col}{hrow + len(rows)}"   # sortable/filterable
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    # ---- RunLog sheet (one row appended per run) ----
    if "RunLog" not in wb.sheetnames:
        rl = wb.create_sheet("RunLog")
        for c, label in enumerate(
            ["run_ts", "space", "region", "threat", "shown", "recorded_rows",
             "snapshots_in_window", "blendable_systems", "best_system", "best_score"], start=1):
            cell = rl.cell(row=1, column=c, value=label)
            cell.font = head_font; cell.fill = head_fill
        rl.freeze_panes = "A2"
        for c, w in enumerate([22, 8, 12, 7, 7, 14, 18, 17, 16, 10], start=1):
            rl.column_dimensions[get_column_letter(c)].width = w
    else:
        rl = wb["RunLog"]
    best = rows[0] if rows else {}
    rl.append([
        meta["ts"], meta["space"], meta["region"] or "all",
        "yes" if meta["threat"] else "no", len(rows), meta["recorded_rows"],
        meta["snapshots_in_window"], meta["blendable_systems"],
        best.get("name", ""), best.get("score", ""),
    ])

    # ---- atomic, Excel-lock-tolerant save ----
    d = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=d)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, path)
        return path
    except PermissionError:
        # target is open in Excel; drop a timestamped side-file instead of failing
        stamp = meta["ts"].replace(":", "").replace("-", "").replace(".", "")[:15]
        alt = os.path.join(d, f"lowsec_scout.{stamp}.xlsx")
        try:
            os.replace(tmp, alt)
        except Exception:
            try: os.remove(tmp)
            except OSError: pass
            print(f"  (Could not write {os.path.basename(path)} -- is it open in Excel? Skipped.)")
            return None
        print(f"  ({os.path.basename(path)} is open in Excel; wrote {os.path.basename(alt)} instead.)")
        return alt
    except Exception as e:
        try: os.remove(tmp)
        except OSError: pass
        print(f"  (xlsx write failed: {e})")
        return None


# ----------------------------------------------------------------------------
# DEMO (offline self-test: exercises cache, recorder, history, scoring, xlsx)
# ----------------------------------------------------------------------------

def demo_cache():
    import random; random.seed(7)
    c = {}
    for i in range(1, 401):
        true = round(random.uniform(0.1, 0.44), 4)   # all lowsec for the demo
        c[str(i)] = {"name": f"Sys-{i}", "sec": round(true, 2), "truesec": true,
                     "region_id": "x", "region": "DemoRegion"}
    return c

def demo_activity(cache):
    import random; random.seed(11)
    k, j = {}, {}
    for sid in list(cache):
        if random.random() < 0.3:
            k[sid] = (random.randint(0, 6), random.randint(0, 4), random.randint(0, 30))
        j[sid] = random.randint(0, 40)
    return k, j

def demo_seed_history(path, cache):
    """Write a few fake past snapshots to the demo history CSV so the averaging
    path runs offline (exercises the real history_load code)."""
    import random; random.seed(3)
    now = datetime.now(timezone.utc)
    rows = []
    for hrs in (1, 2, 3, 26, 50):                       # a few hours + a couple days back
        ts = (now - timedelta(hours=hrs)).isoformat()
        for sid in list(cache)[:120]:
            if random.random() < 0.4:
                rows.append((ts, sid, random.randint(0, 5), random.randint(0, 3),
                             random.randint(0, 20), random.randint(0, 30)))
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(HIST_COLS); w.writerows(rows)


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Rank safe, low-traffic, ore-suitable EVE systems.")
    ap.add_argument("--top", type=int, default=25)
    ap.add_argument("--space", choices=["lowsec", "hisec", "null", "all"], default="lowsec")
    ap.add_argument("--region", default=None, help="limit to one region (name)")
    ap.add_argument("--min-jumps", type=int, default=0, help="exclude systems below this traffic")
    ap.add_argument("--history-days", type=int, default=7, help="lookback window for recorded danger")
    ap.add_argument("--retain-days", type=int, default=120, help="prune history CSV older than this (0=keep all)")
    ap.add_argument("--no-record", action="store_true", help="don't append this run to the history CSV")
    ap.add_argument("--no-threat", action="store_true", help="skip zKill shortlist lookups")
    ap.add_argument("--threat-top", type=int, default=None, help="how many to zKill-check (default=--top)")
    ap.add_argument("--out", default=None, help="also write the shown table to this CSV (one-off)")
    ap.add_argument("--history", default=HIST_CSV, help="running history CSV path")
    ap.add_argument("--xlsx", default=XLSX_PATH, help="running workbook path")
    ap.add_argument("--no-xlsx", action="store_true", help="don't write the .xlsx workbook")
    ap.add_argument("--html", default="index.html", help="write the page here each run (served as the site index)")
    ap.add_argument("--no-html", action="store_true", help="don't write the local HTML page")
    ap.add_argument("--page-top", type=int, default=10, help="how many systems on the page / webhook")
    ap.add_argument("--open", action="store_true", help="open the HTML page in your browser after writing")
    ap.add_argument("--webhook", default=None, help="also POST the page payload to this URL (optional http server)")
    ap.add_argument("--webhook-token", default=os.environ.get("SCOUT_WEBHOOK_TOKEN"),
                    help="optional shared secret sent as X-Webhook-Token")
    ap.add_argument("--demo", action="store_true")
    a = ap.parse_args()

    tag = "DEMO/FAKE" if a.demo else f"LIVE ESI . {a.space}"
    print(f"\n  CROWN & OAK - Lowsec Scout v3  [{tag}]")

    # ---- system map ----
    try:
        cache = demo_cache() if a.demo else build_system_cache()
    except Exception as e:
        print(f"  Could not build system cache: {e}")
        print("  (Check internet / that fuzzwork.co.uk is reachable.)"); return 1

    # ---- live activity ----
    try:
        kills, jumps = demo_activity(cache) if a.demo else get_live_activity()
    except urllib.error.HTTPError as e:
        print(f"  ESI error: HTTP {e.code} ({e.headers.get('x-deny-reason','')})"); return 1
    except Exception as e:
        print(f"  Could not fetch live activity: {e}"); return 1

    # ---- record + load history (the running CSV) ----
    hist_file = "scout_history_demo.csv" if a.demo else a.history
    if a.demo:
        demo_seed_history(hist_file, cache)
    ts = datetime.now(timezone.utc).isoformat()
    recorded = 0
    if not a.no_record:
        recorded = history_record(hist_file, ts, kills, jumps)
        print(f"  Recorded {recorded} system rows to {hist_file}")
        if a.retain_days and a.retain_days > 0:
            try:
                dropped = history_prune(hist_file, a.retain_days)
                if dropped:
                    print(f"  Pruned {dropped} rows older than {a.retain_days} days.")
            except Exception as e:
                print(f"  (history prune skipped: {e})")
    hist, snaps_in_window = history_load(hist_file, a.history_days)

    # ---- score every candidate ----
    rows = []
    for sid, meta in cache.items():
        if not in_band(meta["sec"], a.space): continue
        if a.region and meta["region"].lower() != a.region.lower(): continue
        sk, pk, nk = kills.get(sid, (0, 0, 0))
        jp = jumps.get(sid, 0)
        if jp < a.min_jumps: continue

        h_ship, h_pod, samples, h_max = hist.get(sid, (0, 0, 0, 0))
        live_dw = danger_weighted(sk, pk)
        if samples >= 3:                                  # trust history once we have it
            hist_dw = danger_weighted(h_ship, h_pod)
            dw = 0.6 * hist_dw + 0.4 * live_dw
        else:
            dw = live_dw
        safety = safety_score(dw)
        ore = round(ore_potential(meta["truesec"]))
        rows.append({
            "name": meta["name"], "sec": meta["sec"], "truesec": meta["truesec"],
            "region": meta["region"], "ship": sk, "pod": pk, "npc": nk, "jumps": jp,
            "hist_avg_pvp": round(h_ship + h_pod, 2), "samples": samples,
            "safety": safety, "ore": ore, "score": composite(safety, ore),
            "zk_recent": "", "zk_last_h": "",
        })

    # best first: score desc, then richer ore, then quietest traffic. Ore breaks the
    # very common safety=100 tie so accessibility is not penalised; jumps only breaks
    # ore ties. (This is presentation ordering, not the protected scoring formula.)
    rows.sort(key=lambda r: (-r["score"], -r["ore"], r["jumps"]))
    rows = rows[:a.top]
    if not rows:
        print("  No systems matched."); return 0

    # ---- threat intel on the shortlist ----
    if not a.no_threat and not a.demo:
        n = a.threat_top if a.threat_top is not None else len(rows)
        check = rows[: max(0, n)]   # None -> whole shortlist; 0 -> none; clamp negatives
        print(f"  Cross-checking {len(check)} systems on zKillboard (recent kills)...")
        name_to_id = {m["name"]: sid for sid, m in cache.items()}
        for r in check:
            cnt, last_h = zkill_recent(name_to_id.get(r["name"]))
            r["zk_recent"] = "" if cnt is None else cnt
            r["zk_last_h"] = "" if last_h is None else last_h
    elif a.demo:
        import random; random.seed(5)
        for r in rows:
            r["zk_recent"] = random.randint(0, 30)
            r["zk_last_h"] = round(random.uniform(0.5, 200), 1)

    # ---- print ----
    hdr = (f"  {'SYSTEM':<13}{'SEC':>5} {'REGION':<16}{'SCORE':>6}{'SAFE':>5}{'ORE':>5}"
           f"{'PVP1h':>6}{'JUMP':>5}{'~AVG':>6}{'N':>4}{'ZK':>5}{'LASTh':>7}")
    print("\n" + hdr); print("  " + "-" * (len(hdr) - 2))
    for r in rows:
        print(f"  {r['name'][:12]:<13}{r['sec']:>5.1f} {r['region'][:15]:<16}"
              f"{r['score']:>6}{r['safety']:>5}{r['ore']:>5}{r['ship']+r['pod']:>6}"
              f"{r['jumps']:>5}{r['hist_avg_pvp']:>6}{r['samples']:>4}"
              f"{str(r['zk_recent']):>5}{str(r['zk_last_h']):>7}")

    print("\n  SCORE = 70% safety + 30% ore-potential.  SAFE = low kills/traffic (history-blended).")
    print("  ORE = truesec-based potential (lower sec = richer); PVP1h = ship+pod kills this hour.")
    print("  ~AVG = recorded avg PVP/hr over the last", a.history_days, "days (needs prior runs).")
    print("  N = history samples; once N>=3 the score blends recorded danger, else it is live-only.")
    print("  ZK = recent kills on zKill, LASTh = hours since the last one (longer-window danger).")
    print("  Ties (very common at SAFE=100) break by richer ORE, then lowest traffic.")

    # ---- one-off CSV of the shown table ----
    if a.out:
        cols = ["name", "sec", "truesec", "region", "score", "safety", "ore",
                "ship", "pod", "npc", "jumps", "hist_avg_pvp", "samples",
                "zk_recent", "zk_last_h"]
        with open(a.out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader(); w.writerows(rows)
        print(f"\n  Wrote {len(rows)} rows to {a.out}")

    # ---- running workbook ----
    if not a.no_xlsx:
        xlsx_file = "lowsec_scout_demo.xlsx" if a.demo else a.xlsx
        meta = {"ts": ts, "space": a.space, "region": a.region,
                "history_days": a.history_days, "recorded_rows": recorded,
                "snapshots_in_window": snaps_in_window,
                "blendable_systems": sum(1 for v in hist.values() if v[2] >= 3),
                "threat": (a.demo or not a.no_threat)}
        written = write_xlsx(xlsx_file, rows, meta)
        if written:
            print(f"  Updated workbook {written}  (Shortlist refreshed, RunLog appended)")

    # ---- local page (default) + optional webhook, from one payload ----
    if (not a.no_html) or a.webhook:
        topn = rows[: max(0, a.page_top)]
        state = {
            "generated_at": ts, "space": a.space, "region": a.region,
            "threat": (a.demo or not a.no_threat), "snapshots_in_window": snaps_in_window,
            "systems": [{
                "rank": i + 1, "name": r["name"],
                "dotlan": "https://evemaps.dotlan.net/system/" + str(r["name"]).replace(" ", "_"),
                "sec": r["sec"], "region": r["region"], "score": r["score"],
                "safety": r["safety"], "ore": r["ore"], "pvp1h": r["ship"] + r["pod"],
                "jumps": r["jumps"], "hist_avg_pvp": r["hist_avg_pvp"], "samples": r["samples"],
                "zk_recent": r["zk_recent"], "zk_last_h": r["zk_last_h"], "truesec": r["truesec"],
            } for i, r in enumerate(topn)],
        }
        if not a.no_html:
            html_file = "lowsec_scout_demo.html" if a.demo else a.html
            try:
                from scout_page import write_html
                pw = os.environ.get("EVE_PAGE_PASSWORD")
                write_html(html_file, state, pw)
                lock = "locked (EVE_PAGE_PASSWORD)" if pw else "UNLOCKED (set EVE_PAGE_PASSWORD to lock)"
                print(f"  Wrote page {html_file}  [{lock}]")
                if a.open:
                    import webbrowser
                    webbrowser.open("file://" + os.path.abspath(html_file))
            except Exception as e:
                print(f"  (could not write HTML page: {e})")
        if a.webhook:
            try:
                code, _ = post_webhook(a.webhook, state, a.webhook_token)
                print(f"  Posted page to {a.webhook} (HTTP {code})")
            except Exception as e:
                print(f"  (webhook post failed: {e} -- is scout_server.py running?)")

    if a.demo:
        print("\n  ^ DEMO DATA IS FAKE (incl. seeded history). Run without --demo for live.\n")
    else:
        print("\n  Reminder: ore ANOMALIES aren't in any API. These are safe CANDIDATES; scout")
        print("  and scan in a throwaway ship. Run hourly so ~AVG and history fill in.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
